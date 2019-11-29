import numpy as np
import cv2
import imutils
from pytesseract import *
import pytesseract
from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import *
import openpyxl
img=cv2.imread("car10.jpg", 1)
img=cv2.resize(img,(620,480))
#cv2.imshow('colour image',img)

gray_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
cv2.imwrite('gray_image.png',gray_image)
#cv2.imshow('gray_image',gray_image)


gray = cv2.bilateralFilter(gray_image, 11, 17, 17)
#cv2.imshow('gray_image1',gray)

edged = cv2.Canny(gray, 30, 200)
#cv2.imshow('edges',edged)

nts = cv2.findContours(edged.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
cnts = imutils.grab_contours(nts)
cnts = sorted(cnts, key = cv2.contourArea, reverse = True)[:10]
screenCnt = None
for c in cnts:
                # approximate the contour
                peri = cv2.arcLength(c, True)
                approx = cv2.approxPolyDP(c, 0.018 * peri, True)
                # if our approximated contour has four points, then
                # we can assume that we have found our screen
                if len(approx) == 4:
                      screenCnt = approx
                      break
mask = np.zeros(gray.shape,np.uint8)
new_image = cv2.drawContours(mask,[screenCnt],0,255,-1,)
new_image = cv2.bitwise_and(img,img,mask=mask)
#cv2.imshow('mask',new_image)
cv2.imwrite('ocr.png',new_image)
pytesseract.pytesseract.tesseract_cmd='C:\\Program Files\\Tesseract-OCR\\tesseract.exe'
text=pytesseract.image_to_string(Image.open("ocr.png"),lang='eng')
print(text)
print("CAR NUMBER","\t\tNAME","\tPHONE NUMBER","\tEMAIL ID")

fp="C:\\Users\\Siddhi\\Desktop\\anpd.xlsx"
wb=openpyxl.load_workbook(fp)
sheet=wb.active
max_row = sheet.max_row
max_column = sheet.max_column
flag=0
d=[["MH12DE1433)","abc",9798956544,"abc@yahoo.com"],
["NH13.CD 0096","def",8716426262,"def@gmail.com"],
["ET65 EEM","ghi",7946662121,"ghi@yahoo.com"],
["WH 20 EE 7598","jkl",1454666465,"jkl@gmail.com"],
["TN 9SF 2378","mno",5646974174,"mno@gmail.com"],
["KE S2u o4se","stu",9876543210,"stu@rediff.com"],
["DZI7 YXR","vwx",9638520147,"vwx@gmail.com"]

]

for i in range(1, max_row + 1):
    if sheet.cell(row=i,column=1).value==text:
        for j in range(1, max_column+1):
            print(sheet.cell(row=i,column=j).value, end="\t\t")
        flag=1
        break
if flag==0:

    wb1 = openpyxl.Workbook()
    sheet=wb1.active
    name=input("Enter owner's name:")
    pn=input("Enter phone number:")
    id=input("Enter email id:")
    l=[]
    l.append(text)
    l.append(name)
    l.append(pn)
    l.append(id)
    d.append(l)
    for row in d:
        sheet.append(row)
    wb1.save(fp)

cv2.waitKey(0)
cv2.destroyAllWindows()
